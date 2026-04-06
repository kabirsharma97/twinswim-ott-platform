"""
Generates assets/feature_store_template.xlsx on first run.
Called from brandcine_ott_app.py at startup.
"""
from __future__ import annotations

from pathlib import Path

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "assets" / "feature_store_template.xlsx"

# Field definitions: (field_name, category, mandatory, data_type, expected_range, description)
FIELDS = [
    # USER IDENTITY & CONTENT INTERACTION
    ("user_id", "User Identity & Content Interaction", "MANDATORY", "string/int", "Unique per row", "Primary key — unique user identifier"),
    ("content_id", "User Identity & Content Interaction", "GOOD TO HAVE", "string/int", "Any", "Identifier for the content item watched"),
    ("genre", "User Identity & Content Interaction", "GOOD TO HAVE", "string", "e.g. Drama, Action", "Genre of content watched"),
    ("watch_minutes", "User Identity & Content Interaction", "GOOD TO HAVE", "float", "≥0", "Total minutes watched in session"),
    ("completion_rate", "User Identity & Content Interaction", "MANDATORY", "float", "0.0–1.0", "Fraction of content completed"),
    ("early_drop_rate", "User Identity & Content Interaction", "MANDATORY", "float", "0.0–1.0", "Rate of early drop-offs in sessions"),
    ("mid_drop_rate", "User Identity & Content Interaction", "MANDATORY", "float", "0.0–1.0", "Rate of mid-session exits"),
    ("session_start", "User Identity & Content Interaction", "GOOD TO HAVE", "datetime", "Any", "Timestamp when session started"),
    ("device_type", "User Identity & Content Interaction", "MANDATORY", "string", "mobile/tablet/smart_tv/laptop/desktop", "Primary device type for session"),
    ("is_weekend", "User Identity & Content Interaction", "MANDATORY", "bool/int", "0 or 1", "Whether session occurred on a weekend"),
    ("No_of_Pauses", "User Identity & Content Interaction", "GOOD TO HAVE", "int", "≥0", "Number of pauses during session"),
    ("Content_Tags", "User Identity & Content Interaction", "GOOD TO HAVE", "string", "Any", "Tags associated with the content"),
    ("Content_Length", "User Identity & Content Interaction", "GOOD TO HAVE", "float", "≥0", "Length of content in minutes"),
    # VIEWING CONTEXT & DISCOVERY
    ("watch_context", "Viewing Context & Discovery", "MANDATORY", "string", "e.g. solo/family/background", "Context in which content was watched"),
    ("entry_source", "Viewing Context & Discovery", "MANDATORY", "string", "e.g. search/browse/recommendation", "How user found the content"),
    ("rewatch_flag", "Viewing Context & Discovery", "MANDATORY", "bool/int", "0 or 1", "Whether this was a rewatch"),
    ("skip_intro_rate", "Viewing Context & Discovery", "MANDATORY", "float", "0.0–1.0", "Rate at which user skips intros"),
    # SESSION BEHAVIOUR METRICS
    ("session_id", "Session Behaviour Metrics", "GOOD TO HAVE", "string/int", "Unique per session", "Unique session identifier"),
    ("session_duration_min", "Session Behaviour Metrics", "MANDATORY", "float", "≥0", "Total duration of session in minutes"),
    ("num_titles_watched", "Session Behaviour Metrics", "MANDATORY", "int", "≥0", "Number of titles watched per session"),
    ("time_of_day", "Session Behaviour Metrics", "MANDATORY", "string/int", "e.g. morning/afternoon/evening/night", "Time period when session occurred"),
    ("buffer_events", "Session Behaviour Metrics", "MANDATORY", "int", "≥0", "Number of buffering events in session"),
    ("avg_bitrate_mbps", "Session Behaviour Metrics", "MANDATORY", "float", "≥0", "Average streaming bitrate in Mbps"),
    ("volume_change_events", "Session Behaviour Metrics", "GOOD TO HAVE", "int", "≥0", "Number of volume adjustments"),
    ("pause_cluster_pattern", "Session Behaviour Metrics", "MANDATORY", "string/int", "Any", "Encoded pattern of pause behaviour"),
    ("attention_decay_curve", "Session Behaviour Metrics", "GOOD TO HAVE", "float", "Any", "Measure of attention decay over session"),
    # CONTENT AFFINITY & BEHAVIORAL TRAITS
    ("fav_genre", "Content Affinity & Behavioral Traits", "MANDATORY", "string", "e.g. Drama, Action", "User's favourite genre"),
    ("sports_dependency_score", "Content Affinity & Behavioral Traits", "MANDATORY", "float", "0.0–1.0", "Propensity to watch sports content"),
    ("binge_index", "Content Affinity & Behavioral Traits", "MANDATORY", "float", "0.0–1.0", "Binge-watching propensity score"),
    ("avg_watch_gap_days", "Content Affinity & Behavioral Traits", "MANDATORY", "float", "≥0", "Average days between viewing sessions"),
    # EDITORIAL / DEVICE CONTEXT
    ("secondary_device", "Editorial / Device Context", "GOOD TO HAVE", "string", "Any", "Secondary device type used"),
    ("smart_tv_brand", "Editorial / Device Context", "GOOD TO HAVE", "string", "Any", "Smart TV brand if applicable"),
    ("app_version", "Editorial / Device Context", "GOOD TO HAVE", "string", "Any", "App version in use"),
    ("screen_size_category", "Editorial / Device Context", "GOOD TO HAVE", "string", "small/medium/large", "Screen size category"),
    ("casting_usage_flag", "Editorial / Device Context", "GOOD TO HAVE", "bool/int", "0 or 1", "Whether casting was used"),
    ("hdr_support_flag", "Editorial / Device Context", "GOOD TO HAVE", "bool/int", "0 or 1", "Whether HDR is supported"),
    ("os_version_major", "Editorial / Device Context", "GOOD TO HAVE", "int", "Any", "Major OS version number"),
    # SUBSCRIPTION & COMMERCIAL BEHAVIOUR
    ("plan_type", "Subscription & Commercial Behaviour", "MANDATORY", "string", "basic/standard/premium", "Subscription plan type"),
    ("monthly_price", "Subscription & Commercial Behaviour", "MANDATORY", "float", "≥0", "Monthly subscription price"),
    ("is_bundle", "Subscription & Commercial Behaviour", "MANDATORY", "bool/int", "0 or 1", "Whether on a bundle plan"),
    ("tenure_months", "Subscription & Commercial Behaviour", "MANDATORY", "int", "1–60", "Months as a subscriber"),
    ("discount_flag", "Subscription & Commercial Behaviour", "MANDATORY", "bool/int", "0 or 1", "Whether discount is applied"),
    ("discount_expectation_flag", "Subscription & Commercial Behaviour", "MANDATORY", "bool/int", "0 or 1", "Whether user expects a discount"),
    # RISK, CHURN & PIRACY BEHAVIOUR
    ("churn_probability", "Risk, Churn & Piracy Behaviour", "MANDATORY", "float", "0.0–1.0", "Model-derived churn probability"),
    ("last_active_days_ago", "Risk, Churn & Piracy Behaviour", "MANDATORY", "int", "≥0", "Days since last active session"),
    ("piracy_exposure_flag", "Risk, Churn & Piracy Behaviour", "MANDATORY", "bool", "True/False", "Whether user has piracy exposure"),
    ("payment_failure_count", "Risk, Churn & Piracy Behaviour", "GOOD TO HAVE", "int", "≥0", "Number of payment failures"),
    ("piracy_trigger_reason", "Risk, Churn & Piracy Behaviour", "GOOD TO HAVE", "string", "Any (structural nulls OK)", "Reason for piracy trigger — null for non-piracy users"),
    ("content_unavailable_flag", "Risk, Churn & Piracy Behaviour", "MANDATORY", "bool/int", "0 or 1", "Content unavailability experienced"),
    ("vpn_usage_suspected", "Risk, Churn & Piracy Behaviour", "MANDATORY", "bool/int", "0 or 1", "VPN usage suspected"),
    ("piracy_recency_score", "Risk, Churn & Piracy Behaviour", "MANDATORY", "float", "0.0–1.0", "Recency score of piracy-related behaviour"),
    # SUPPORT INTERACTION & TRUST
    ("ticket_id", "Support Interaction & Trust", "GOOD TO HAVE", "string/int", "Any", "Support ticket identifier"),
    ("issue_type", "Support Interaction & Trust", "GOOD TO HAVE", "string", "Any", "Category of support issue"),
    ("resolution_time_hrs", "Support Interaction & Trust", "GOOD TO HAVE", "float", "≥0", "Hours taken to resolve support ticket"),
    ("csat_score", "Support Interaction & Trust", "GOOD TO HAVE", "float", "1–5 or 1–10", "Customer satisfaction score"),
    ("support_channel", "Support Interaction & Trust", "GOOD TO HAVE", "string", "chat/email/phone", "Channel used for support"),
    ("issue_repeat_flag", "Support Interaction & Trust", "GOOD TO HAVE", "bool/int", "0 or 1", "Whether issue was a repeat"),
    # CUSTOMER SENTIMENT & ADVOCACY
    ("nps_score", "Customer Sentiment & Advocacy", "MANDATORY", "int/float", "0–10", "Net Promoter Score"),
    ("content_satisfaction", "Customer Sentiment & Advocacy", "MANDATORY", "float", "1–5 or 1–10", "Content satisfaction rating"),
    ("price_perception", "Customer Sentiment & Advocacy", "MANDATORY", "float", "1–5 or 1–10", "Perceived value for price"),
    # SOCIAL MEDIA SIGNALS
    ("post_id", "Social Media Signals", "GOOD TO HAVE", "string/int", "Any", "Social media post identifier"),
    ("platform", "Social Media Signals", "GOOD TO HAVE", "string", "twitter/instagram/facebook", "Social platform"),
    ("sentiment_score", "Social Media Signals", "GOOD TO HAVE", "float", "-1.0 to 1.0", "Sentiment score of social post"),
    ("mentioned_genre", "Social Media Signals", "GOOD TO HAVE", "string", "Any", "Genre mentioned in social post"),
    # NETWORK & GEOGRAPHIC CONTEXT
    ("city", "Network & Geographic Context", "MANDATORY", "string", "Any", "User's city"),
    ("isp_partner", "Network & Geographic Context", "MANDATORY", "string", "Any", "Internet service provider name"),
    ("network_type", "Network & Geographic Context", "MANDATORY", "string", "4G/5G/Fiber/Broadband", "Network connection type"),
    ("event_calendar_flag", "Network & Geographic Context", "MANDATORY", "bool/int", "0 or 1", "Whether a major event was on during session"),
    ("urbanicity", "Network & Geographic Context", "MANDATORY", "string", "urban/suburban/rural", "Urban classification of user location"),
    ("avg_network_jitter", "Network & Geographic Context", "MANDATORY", "float", "≥0 (ms)", "Average network jitter in milliseconds"),
    ("peak_hour_congestion_flag", "Network & Geographic Context", "MANDATORY", "bool/int", "0 or 1", "Whether session was during peak congestion hours"),
]

# Sample data rows (3 realistic rows)
SAMPLE_ROWS = [
    {
        "user_id": "U100001", "content_id": "C5021", "genre": "Drama", "watch_minutes": 48.5,
        "completion_rate": 0.82, "early_drop_rate": 0.08, "mid_drop_rate": 0.10,
        "session_start": "2024-03-15 20:30:00", "device_type": "smart_tv", "is_weekend": 0,
        "No_of_Pauses": 2, "Content_Tags": "crime|thriller", "Content_Length": 58.0,
        "watch_context": "solo", "entry_source": "recommendation", "rewatch_flag": 0,
        "skip_intro_rate": 0.75, "session_id": "S200001", "session_duration_min": 92.3,
        "num_titles_watched": 2, "time_of_day": "evening", "buffer_events": 1,
        "avg_bitrate_mbps": 8.4, "volume_change_events": 3, "pause_cluster_pattern": "sparse",
        "attention_decay_curve": 0.12, "fav_genre": "Drama", "sports_dependency_score": 0.15,
        "binge_index": 0.68, "avg_watch_gap_days": 2.3, "secondary_device": "mobile",
        "smart_tv_brand": "Samsung", "app_version": "4.2.1", "screen_size_category": "large",
        "casting_usage_flag": 0, "hdr_support_flag": 1, "os_version_major": 13,
        "plan_type": "premium", "monthly_price": 16.99, "is_bundle": 0, "tenure_months": 18,
        "discount_flag": 0, "discount_expectation_flag": 0, "churn_probability": 0.12,
        "last_active_days_ago": 1, "piracy_exposure_flag": False, "payment_failure_count": 0,
        "piracy_trigger_reason": None, "content_unavailable_flag": 0, "vpn_usage_suspected": 0,
        "piracy_recency_score": 0.03, "ticket_id": None, "issue_type": None,
        "resolution_time_hrs": None, "csat_score": None, "support_channel": None,
        "issue_repeat_flag": None, "nps_score": 9, "content_satisfaction": 4.2,
        "price_perception": 3.8, "post_id": None, "platform": None, "sentiment_score": None,
        "mentioned_genre": None, "city": "Mumbai", "isp_partner": "Jio", "network_type": "Fiber",
        "event_calendar_flag": 0, "urbanicity": "urban", "avg_network_jitter": 8.2,
        "peak_hour_congestion_flag": 1,
    },
    {
        "user_id": "U100002", "content_id": "C3312", "genre": "Sports", "watch_minutes": 110.0,
        "completion_rate": 0.95, "early_drop_rate": 0.02, "mid_drop_rate": 0.03,
        "session_start": "2024-03-16 15:00:00", "device_type": "mobile", "is_weekend": 1,
        "No_of_Pauses": 0, "Content_Tags": "cricket|live", "Content_Length": 115.0,
        "watch_context": "family", "entry_source": "search", "rewatch_flag": 0,
        "skip_intro_rate": 0.10, "session_id": "S200002", "session_duration_min": 118.5,
        "num_titles_watched": 1, "time_of_day": "afternoon", "buffer_events": 4,
        "avg_bitrate_mbps": 5.1, "volume_change_events": 1, "pause_cluster_pattern": "none",
        "attention_decay_curve": 0.04, "fav_genre": "Sports", "sports_dependency_score": 0.91,
        "binge_index": 0.21, "avg_watch_gap_days": 0.5, "secondary_device": "smart_tv",
        "smart_tv_brand": None, "app_version": "4.1.9", "screen_size_category": "small",
        "casting_usage_flag": 0, "hdr_support_flag": 0, "os_version_major": 14,
        "plan_type": "standard", "monthly_price": 12.99, "is_bundle": 1, "tenure_months": 6,
        "discount_flag": 1, "discount_expectation_flag": 1, "churn_probability": 0.34,
        "last_active_days_ago": 0, "piracy_exposure_flag": True, "payment_failure_count": 1,
        "piracy_trigger_reason": "geo_restriction", "content_unavailable_flag": 1,
        "vpn_usage_suspected": 1, "piracy_recency_score": 0.55, "ticket_id": "T8821",
        "issue_type": "stream_quality", "resolution_time_hrs": 4.5, "csat_score": 3.0,
        "support_channel": "chat", "issue_repeat_flag": 0, "nps_score": 6,
        "content_satisfaction": 3.5, "price_perception": 2.9, "post_id": "P1122",
        "platform": "twitter", "sentiment_score": -0.32, "mentioned_genre": "Sports",
        "city": "Delhi", "isp_partner": "Airtel", "network_type": "4G",
        "event_calendar_flag": 1, "urbanicity": "urban", "avg_network_jitter": 22.7,
        "peak_hour_congestion_flag": 1,
    },
    {
        "user_id": "U100003", "content_id": "C7890", "genre": "Comedy", "watch_minutes": 22.0,
        "completion_rate": 0.55, "early_drop_rate": 0.30, "mid_drop_rate": 0.15,
        "session_start": "2024-03-14 09:15:00", "device_type": "laptop", "is_weekend": 0,
        "No_of_Pauses": 5, "Content_Tags": "sitcom|family", "Content_Length": 40.0,
        "watch_context": "background", "entry_source": "browse", "rewatch_flag": 1,
        "skip_intro_rate": 0.95, "session_id": "S200003", "session_duration_min": 35.0,
        "num_titles_watched": 3, "time_of_day": "morning", "buffer_events": 0,
        "avg_bitrate_mbps": 12.0, "volume_change_events": 8, "pause_cluster_pattern": "clustered",
        "attention_decay_curve": 0.45, "fav_genre": "Comedy", "sports_dependency_score": 0.05,
        "binge_index": 0.42, "avg_watch_gap_days": 8.1, "secondary_device": None,
        "smart_tv_brand": None, "app_version": "4.3.0", "screen_size_category": "medium",
        "casting_usage_flag": 0, "hdr_support_flag": 0, "os_version_major": 11,
        "plan_type": "basic", "monthly_price": 7.99, "is_bundle": 0, "tenure_months": 3,
        "discount_flag": 0, "discount_expectation_flag": 1, "churn_probability": 0.67,
        "last_active_days_ago": 5, "piracy_exposure_flag": False, "payment_failure_count": 0,
        "piracy_trigger_reason": None, "content_unavailable_flag": 0, "vpn_usage_suspected": 0,
        "piracy_recency_score": 0.08, "ticket_id": None, "issue_type": None,
        "resolution_time_hrs": None, "csat_score": None, "support_channel": None,
        "issue_repeat_flag": None, "nps_score": 7, "content_satisfaction": 3.0,
        "price_perception": 4.1, "post_id": None, "platform": None, "sentiment_score": None,
        "mentioned_genre": None, "city": "Pune", "isp_partner": "BSNL", "network_type": "Broadband",
        "event_calendar_flag": 0, "urbanicity": "suburban", "avg_network_jitter": 15.3,
        "peak_hour_congestion_flag": 0,
    },
]


def create_template(output_path: Path | None = None) -> None:
    """Generate the feature store template Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out = output_path or TEMPLATE_PATH
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feature Store"

    # Styles
    header_fill = PatternFill("solid", fgColor="0D1B2A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    mandatory_fill = PatternFill("solid", fgColor="FFFFFF")
    optional_fill = PatternFill("solid", fgColor="FFF9C4")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["Field Name", "Category", "Mandatory", "Data Type", "Expected Range/Values", "Description"]
    col_widths = [28, 38, 16, 18, 30, 55]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, (fname, category, mandatory, dtype, expected, desc) in enumerate(FIELDS, start=2):
        is_mandatory = mandatory == "MANDATORY"
        fill = mandatory_fill if is_mandatory else optional_fill
        row_data = [fname, category, mandatory, dtype, expected, desc]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = left_wrap
            if col_idx == 3:
                cell.font = Font(bold=True)
                cell.alignment = center

    # Sample Data sheet
    ws2 = wb.create_sheet("Sample Data")
    all_fields = [f[0] for f in FIELDS]

    for col_idx, field in enumerate(all_fields, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18

    ws2.freeze_panes = "A2"
    for row_idx, sample in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, field in enumerate(all_fields, start=1):
            val = sample.get(field, "")
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = left_wrap

    wb.save(str(out))
    print(f"Feature store template created → {out}")


def ensure_template() -> None:
    """Create template only if it doesn't already exist."""
    if not TEMPLATE_PATH.exists():
        try:
            create_template()
        except Exception as e:
            print(f"Warning: could not create feature store template: {e}")


if __name__ == "__main__":
    create_template()
